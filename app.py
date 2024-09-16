from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import io
from io import BytesIO
import pdfplumber
import pandas as pd
from functools import wraps
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import openpyxl

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Change this to a more secure key

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

login_credentials = {
    'Admin': generate_password_hash('password123'),
    'user': generate_password_hash('password456'),
    'Adithya' :generate_password_hash("iniya"),
    'JenitaMaam' : generate_password_hash("Professor")
}

subject_credits = {}  # Define subject_credits globally
update_subject_credits = {}
primary_file = None
secondary_file = None

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_update_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def allowed_file(filename, extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in first.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in login_credentials and check_password_hash(login_credentials[username], password):
            session['username'] = username
            return redirect(url_for('home'))  # Redirect to 'home' instead of 'index'
        else:
            flash('Invalid username or password.', 'error')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/home')
@login_required
def home():
    return render_template('home.html')

@app.route('/analyse')
@login_required
def analyse():
    return render_template('analyse.html')

@app.route('/update_subjects')
def update_subjects():
    return render_template('update_subjects.html', subjects=update_subject_credits)

@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/model_analysis')
def model_analysis():
    return render_template('model_analysis.html')


@app.route('/add_update_subject', methods=['POST'])
def add_update_subject():
    subject_code = request.form['subject_code'].strip()
    credits = request.form['credits'].strip()
    
    if subject_code and credits.isdigit():
        update_subject_credits[subject_code] = int(credits)
        flash('Subject added successfully.', 'success')
    else:
        flash('Subject code and credits are required.', 'danger')
    
    return redirect(url_for('update_subjects'))

@app.route('/upload_files', methods=['GET', 'POST'])
def upload_files():
    global primary_file, secondary_file
    
    if request.method == 'POST':
        if 'primary_file' not in request.files or 'secondary_file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
        
        primary_file = request.files['primary_file']
        secondary_file = request.files['secondary_file']
        color = request.form['color']
        
        if primary_file.filename == '' or secondary_file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
        
        if primary_file and allowed_update_file(primary_file.filename) and secondary_file and allowed_update_file(secondary_file.filename):
            primary_filename = secure_filename(primary_file.filename)
            secondary_filename = secure_filename(secondary_file.filename)
            primary_path = os.path.join(app.config['UPLOAD_FOLDER'], primary_filename)
            secondary_path = os.path.join(app.config['UPLOAD_FOLDER'], secondary_filename)
            primary_file.save(primary_path)
            secondary_file.save(secondary_path)
            
            flash('Files uploaded successfully.', 'success')
            return redirect(url_for('process_files', primary_file=primary_filename, secondary_file=secondary_filename, color=color))
        else:
            flash('Allowed file types are xlsx', 'danger')
            return redirect(request.url)
    
    return render_template('upload_files.html')

@app.route('/process_files')
def process_files():
    primary_filename = request.args.get('primary_file')
    secondary_filename = request.args.get('secondary_file')
    color = request.args.get('color')
    primary_path = os.path.join(app.config['UPLOAD_FOLDER'], primary_filename)
    secondary_path = os.path.join(app.config['UPLOAD_FOLDER'], secondary_filename)
    
    primary_wb = load_workbook(filename=primary_path)
    primary_ws = primary_wb.active
    secondary_wb = load_workbook(filename=secondary_path)
    secondary_ws = secondary_wb.active
    
    delete_columns_with_repeating_value(primary_ws)
    delete_columns_with_repeating_value(secondary_ws)
    
    updated_wb = update_primary_file(primary_ws, secondary_ws, color)
    if updated_wb:
        gpa_wb = calculate_update_gpa(updated_wb)
        output = io.BytesIO()
        gpa_wb.save(output)
        output.seek(0)
        
        return send_file(output, download_name="updated_grades.xlsx", as_attachment=True)
    
    flash('An error occurred while processing the files.', 'danger')
    return redirect(url_for('upload_files'))


@app.route('/upload', methods=['POST'])
@login_required
def upload():
    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(request.url)
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(request.url)
    
    if file and allowed_file(file.filename, ALLOWED_EXTENSIONS):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        
        if filename.endswith('.xlsx'):
            return redirect(url_for('add_subjects', filename=filename))
        elif filename.endswith('.pdf'):
            excel_file_path = extract_tables(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            if excel_file_path:
                return send_file(excel_file_path, as_attachment=True)
            else:
                flash('No tables found in the PDF file.', 'error')
                return redirect(url_for('index'))
        
    else:
        flash('Invalid file format', 'error')
    
    return redirect(request.url)

@app.route('/add_subjects/<filename>', methods=['GET', 'POST'])
@login_required
def add_subjects(filename):
    global subject_credits
    if request.method == 'POST':
        if 'add' in request.form:
            subject_code = request.form.get("subject_code")
            credits = request.form.get("credits")
            if subject_code and credits:
                subject_credits[subject_code] = int(credits)
                flash("Subject added successfully.", "success")
            else:
                flash("Subject code and credits are required.", "error")

        if 'submit' in request.form:
            excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            try:
                wb = load_workbook(filename=excel_file_path)
                ws = wb.active

                filter_rows(ws)
                reset_serial_numbers(ws)
                calculate_gpa(ws)

                wb.save(excel_file_path)

                return send_file(excel_file_path, as_attachment=True, download_name='processed.xlsx')
            except Exception as e:
                flash(f"Error processing file: {str(e)}", "error")
                return redirect(url_for('index'))

    return render_template('add_subjects.html', filename=filename, subject_credits=subject_credits)

@app.route('/calculate_gpa/<filename>', methods=['GET'])
@login_required
def calculate_gpa_route(filename):
    calculate_gpa(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    return redirect(url_for('index'))

@app.route('/pdf_converter', methods=['GET', 'POST'])
@login_required
def pdf_converter():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename, {'pdf'}):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            excel_file_path = extract_tables(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            if excel_file_path:
                return send_file(excel_file_path, as_attachment=True, download_name='output.xlsx')
            else:
                flash('No tables found in the PDF file.', 'error')
                return redirect(url_for('pdf_converter'))
        
        else:
            flash('Invalid file format', 'error')
        
        return redirect(request.url)
    
    return render_template('pdf_converter.html')

def delete_columns_with_repeating_value(ws):
    first_row_values = []
    repeating_column_indices = []

    for col in ws.iter_cols(min_row=1, max_row=1):
        cell_value = col[0].value
        if cell_value is None:
            continue
        if cell_value in first_row_values:
            repeating_column_indices.append(col[0].column)
        first_row_values.append(cell_value)

    if repeating_column_indices:
        second_repeating_index = repeating_column_indices[0]
        while ws.max_column >= second_repeating_index:
            ws.delete_cols(second_repeating_index)

def update_primary_file(primary_ws, secondary_ws, color):
    try:
        for row_idx in range(2, primary_ws.max_row + 1):
            primary_value = primary_ws.cell(row=row_idx, column=2).value
            
            for secondary_row_idx in range(2, secondary_ws.max_row + 1):
                secondary_value = secondary_ws.cell(row=secondary_row_idx, column=2).value
                
                if primary_value == secondary_value:
                    for col_idx in range(4, primary_ws.max_column + 1):
                        primary_header_value = primary_ws.cell(row=1, column=col_idx).value
                        secondary_header_value = secondary_ws.cell(row=1, column=col_idx).value
                        
                        if primary_header_value == secondary_header_value:
                            primary_cell_value = primary_ws.cell(row=row_idx, column=col_idx).value
                            secondary_cell_value = secondary_ws.cell(row=secondary_row_idx, column=col_idx).value
                            
                            if secondary_cell_value != '-' and secondary_cell_value is not None:
                                primary_ws.cell(row=row_idx, column=col_idx).value = secondary_cell_value
                                primary_ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type='solid')
        
        return primary_ws.parent
    except Exception as e:
        flash(f"An error occurred while updating the primary file: {e}", 'danger')
        return None


def extract_tables(file_path):
    tables = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables()
            if page_tables:
                tables.extend(page_tables)

    if tables:
        combined_tables = []
        current_table = []
        for table in tables:
            current_table.extend(table)
            num_rows = len(current_table)
            if num_rows <= 3 and current_table[0][0] != "":
                if len(combined_tables) > 0:
                    combined_tables[-1].extend(current_table)
                else:
                    combined_tables.append(current_table)
                current_table = []
            elif num_rows > 5:
                combined_tables.append(current_table)
                current_table = []

        final_combined_table = []
        for table in combined_tables:
            if not final_combined_table:
                final_combined_table.extend(table)
            else:
                found_duplicate = False
                for prev_table in final_combined_table:
                    if table[:3][-1] == prev_table[:3][-1]:
                        found_duplicate = True
                        break
                if not found_duplicate:
                    final_combined_table.extend(table)

        final_headings = []
        for row in final_combined_table:
            if row not in final_headings:
                final_headings.append(row)

        df = pd.DataFrame(final_headings)

        excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')
        df.to_excel(excel_file_path, index=False)
        return excel_file_path
    else:
        return None
    
def calculate_update_gpa(wb):
    ws = wb.active
    
    filter_rows(ws)
    reset_serial_numbers(ws)
    
    max_row_index = ws.max_row
    max_column_index = ws.max_column
    
    if max_row_index >= 2 and max_column_index >= 4:
        new_column_index = max_column_index + 1
        
        for col_idx in range(4, max_column_index + 1):
            ws.cell(row=1, column=new_column_index + col_idx - 4).value = ws.cell(row=1, column=col_idx).value
        
        for col_idx in range(4, max_column_index + 1):
            for row_idx in range(2, max_row_index + 1):
                grade_value = ws.cell(row=row_idx, column=col_idx).value
                grade_point = 0
                
                if grade_value == "O":
                    grade_point = 10
                elif grade_value == "A+":
                    grade_point = 9
                elif grade_value == "A":
                    grade_point = 8
                elif grade_value == "B+":
                    grade_point = 7
                elif grade_value == "B":
                    grade_point = 6
                elif grade_value == "C":
                    grade_point = 5
                elif grade_value == "U":
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFFFCCCB", end_color="FFFFCCCB", fill_type="solid")
                    grade_point = 0
                elif grade_value == "AB":
                    continue
                else:
                    continue
                
                ws.cell(row=row_idx, column=new_column_index + col_idx - 4).value = grade_point
        
        max_column_index = ws.max_column
        absent_col_index = max_column_index + 1
        fails_col_index = max_column_index + 2
        
        ws.cell(row=1, column=absent_col_index).value = "Absent"
        ws.cell(row=1, column=fails_col_index).value = "Fails"
        
        for row_idx in range(2, max_row_index + 1):
            num_absent = 0
            num_fails = 0
            
            for col_idx in range(4, max_column_index + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                
                if cell_value == "AB":
                    num_absent += 1
                
                if cell_value == "U":
                    num_fails += 1
            
            ws.cell(row=row_idx, column=absent_col_index).value = num_absent
            ws.cell(row=row_idx, column=fails_col_index).value = num_fails
        
        apply_styles(ws)
        
        total_credits_col_index = max_column_index + 3
        gpa_col_index = max_column_index + 4
        
        ws.cell(row=1, column=total_credits_col_index).value = "Total Credits"
        ws.cell(row=1, column=gpa_col_index).value = "GPA"
        
        for row_idx in range(2, max_row_index + 1):
            total_credits = 0
            total_points = 0
            
            for col_idx in range(4, max_column_index + 1):
                subject_code = ws.cell(row=1, column=col_idx).value
                
                if subject_code in update_subject_credits:
                    grade_point = ws.cell(row=row_idx, column=col_idx).value
                    
                    try:
                        grade_point = int(grade_point)
                    except (ValueError, TypeError):
                        grade_point = 0
                    
                    credits = update_subject_credits[subject_code]
                    total_credits += credits
                    total_points += credits * grade_point
            
            gpa = total_points / total_credits if total_credits != 0 else 0
            
            ws.cell(row=row_idx, column=total_credits_col_index).value = total_points
            ws.cell(row=row_idx, column=gpa_col_index).value = gpa*2
        
        apply_styles(ws)
    
    return wb

def calculate_gpa(ws):
    try:
        print("Starting filter_rows")
        filter_rows(ws)
        print("Completed filter_rows")

        print("Starting reset_serial_numbers")
        reset_serial_numbers(ws)
        print("Completed reset_serial_numbers")

        max_row_index = ws.max_row
        max_column_index = ws.max_column

        if max_row_index >= 2 and max_column_index >= 4:
            new_column_index = max_column_index + 1

            # Adding new columns for grade points
            for col_idx in range(4, max_column_index + 1):
                ws.cell(row=1, column=new_column_index + col_idx - 4).value = ws.cell(row=1, column=col_idx).value

            # Calculate grade points
            for col_idx in range(4, max_column_index + 1):
                for row_idx in range(2, max_row_index + 1):
                    grade_value = ws.cell(row=row_idx, column=col_idx).value
                    grade_point = 0

                    if grade_value == "O":
                        grade_point = 10
                    elif grade_value == "A+":
                        grade_point = 9
                    elif grade_value == "A":
                        grade_point = 8
                    elif grade_value == "B+":
                        grade_point = 7
                    elif grade_value == "B":
                        grade_point = 6
                    elif grade_value == "C":
                        grade_point = 5
                    elif grade_value == "U":
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFFFCCCB", end_color="FFFFCCCB", fill_type="solid")
                        grade_point = 0
                    elif grade_value == "AB":
                        continue

                    ws.cell(row=row_idx, column=new_column_index + col_idx - 4).value = grade_point

            # Add columns for "Absent" and "Fails"
            max_column_index = ws.max_column
            absent_col_index = max_column_index + 1
            fails_col_index = max_column_index + 2

            ws.cell(row=1, column=absent_col_index).value = "Absent"
            ws.cell(row=1, column=fails_col_index).value = "Fails"

            # Calculate Absent and Fails counts
            for row_idx in range(2, max_row_index + 1):
                num_absent = 0
                num_fails = 0

                for col_idx in range(4, max_column_index + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value

                    if cell_value == "AB":
                        num_absent += 1

                    if cell_value == "U":
                        num_fails += 1

                ws.cell(row=row_idx, column=absent_col_index).value = num_absent
                ws.cell(row=row_idx, column=fails_col_index).value = num_fails

            # Add columns for "Total Credits" and "GPA"
            total_credits_col_index = max_column_index + 3
            gpa_col_index = max_column_index + 4

            ws.cell(row=1, column=total_credits_col_index).value = "Total Credits"
            ws.cell(row=1, column=gpa_col_index).value = "GPA"

            # Calculate Total Credits and GPA
            for row_idx in range(2, max_row_index + 1):
                total_credits = 0
                total_points = 0

                for col_idx in range(4, max_column_index + 1):
                    subject_code = ws.cell(row=1, column=col_idx).value

                    if subject_code in subject_credits:
                        grade_point = ws.cell(row=row_idx, column=col_idx).value

                        try:
                            grade_point = int(grade_point)
                        except (ValueError, TypeError):
                            grade_point = 0

                        credits = subject_credits[subject_code]
                        total_credits += credits * grade_point
                        total_points += credits 

                gpa = total_credits / total_points if total_credits != 0 else 0

                ws.cell(row=row_idx, column=total_credits_col_index).value = total_credits
                ws.cell(row=row_idx, column=gpa_col_index).value = gpa*2
                print(total_credits,total_points,gpa)

        print("Starting apply_styles")
        apply_styles(ws)
        print("Completed apply_styles")

        flash("GPA calculation completed and styles applied.", "success")
    except Exception as e:
        flash(f"Error processing worksheet: {str(e)}", "error")

def apply_styles(ws):
    # Define the border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define the font style
    cambria_font = Font(name='Cambria', size=11)

    # Define the alignment style
    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )

    # Apply styles to all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Apply the defined styles
            cell.border = thin_border
            cell.font = cambria_font
            cell.alignment = center_alignment

def filter_rows(ws):
    rows_to_delete = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True, min_row=2), start=2):
        value = row[1]
        if not str(value).startswith("312322"):
            rows_to_delete.append(row_idx)

    print(f"Rows to delete: {rows_to_delete}")

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    print("Completed filtering rows")

def reset_serial_numbers(ws):
    serial_number = 1
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).value = serial_number
        serial_number += 1

@app.route('/convert_to_excel', methods=['POST'])
def convert_to_excel():
    if 'file' not in request.files:
        return 'No file part'
    
    file = request.files['file']
    
    if file.filename == '':
        return 'No selected file'
    
    if file and file.filename.endswith('.pdf'):
        output_data = []
        first_page_data = []

        with pdfplumber.open(file) as pdf:
            # Extract tables from the first page
            first_page = pdf.pages[0]
            first_page_tables = first_page.extract_tables()
            for table in first_page_tables:
                df = pd.DataFrame(table)
                first_page_data.append(df)
            
            # Extract tables from the second page onwards
            for page_num, page in enumerate(pdf.pages[1:], start=2):
                tables = page.extract_tables()
                for table in tables:
                    df = pd.DataFrame(table)
                    output_data.append(df)

        if not first_page_data:
            return 'No tables found on the first page of the PDF file.'

        if not output_data:
            return 'No tables found from the second page till the last page of the PDF file.'

        # Combine the data for the first page
        first_page_df = pd.concat(first_page_data, ignore_index=True)
        first_page_df.columns = first_page_df.iloc[0]
        first_page_df = first_page_df.drop(0).reset_index(drop=True)
        first_page_df.columns.name = None

        # Combine the data from the second page onwards
        combined_df = pd.concat(output_data, ignore_index=True)
        combined_df.columns = combined_df.iloc[0]
        combined_df = combined_df.drop(0).reset_index(drop=True)
        combined_df.columns.name = None

        # Add the 'ARREAR' column to the combined_df
        combined_df['ARREAR'] = ""  # Initialize the column with empty strings or a default value

        # Convert all applicable values in the second page to numbers
        combined_df = combined_df.apply(pd.to_numeric, errors='ignore')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the first page data to a new sheet
            first_page_df.to_excel(writer, index=False, sheet_name='Subjects Analysis')
            
            # Write the combined data from the rest of the pages
            combined_df.to_excel(writer, index=False, sheet_name='Students Marks')

            # Format numeric columns in the 'Students Marks' sheet
            workbook = writer.book
            worksheet = writer.sheets['Students Marks']

            for column in combined_df.columns:
                # Check if the column contains numeric data
                if pd.api.types.is_numeric_dtype(combined_df[column]):
                    col_idx = combined_df.columns.get_loc(column) + 1  # Get the Excel column index (1-based)
                    # Apply number format with 0 decimals
                    for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in cell:
                            c.number_format = '0'
                    # Adjust column width to fit the numbers
                    max_length = max(combined_df[column].astype(str).map(len).max(), len(str(combined_df.columns[col_idx - 1])))
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

        output.seek(0)
        
        return send_file(output, download_name="Model_Excel.xlsx", as_attachment=True)
    else:
        return 'Invalid file format. Please upload a PDF file.'
    
@app.route('/download_template', methods=['GET'])
def download_template():
    # Assuming you want to provide a static file for download, e.g., a template or report
    file_path = 'assets/Template.docx'  # Adjust the path to your file
    return send_file(file_path, as_attachment=True, download_name="template.docx")


@app.route('/model_analyse', methods=['POST'])
def model_analyse():
    if 'file' not in request.files:
        return 'No file part'
    
    file = request.files['file']
    
    if file.filename == '':
        return 'No selected file'
    
    if file and file.filename.endswith('.xlsx'):
        # Load the Excel file and read both sheets
        xls = pd.ExcelFile(file)
        first_sheet_df = pd.read_excel(xls, sheet_name=0)  # Load the first sheet
        df = pd.read_excel(xls, sheet_name=1)  # Load the second sheet
        
        # Perform operations on the second sheet (df)
        subjects = df.columns[1:].tolist()
        subjects.remove('ARREAR')

        combined_results, failure_counts = calculate_results(df, subjects)

        df['Fail Count'] = failure_counts

        results_df = pd.DataFrame(combined_results).T
        results_df.reset_index(inplace=True)
        results_df.rename(columns={'index': 'Subject'}, inplace=True)

        transposed_results_df = results_df.transpose()
        transposed_results_df.columns = transposed_results_df.iloc[0]
        transposed_results_df = transposed_results_df.drop(transposed_results_df.index[0])

        # Styling setup
        font = Font(name='Times New Roman', size=14)
        bold_font = Font(name='Times New Roman', size=14, bold=True)
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the first sheet with the required styling but no data modification
            first_sheet_df.to_excel(writer, index=False, sheet_name='First Sheet')
            first_sheet_ws = writer.sheets['First Sheet']
            for row in first_sheet_ws.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border
                    cell.alignment = alignment
            for cell in first_sheet_ws[1]:
                cell.font = bold_font

            # Write the modified second sheet
            df.to_excel(writer, index=False, sheet_name='Combined Table')

            worksheet = writer.sheets['Combined Table']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border
                    cell.alignment = alignment

                    # Convert cell to number format with 0 decimals if it's a number
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
                        cell.value = round(float(cell.value))

            # Adjust column widths to fit the formatted numbers
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                adjusted_width = (max_length + 2) * 1.2  # Adding extra space and adjusting factor
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            for cell in worksheet[1]:
                cell.font = bold_font

            # Write the transposed results to a new sheet
            transposed_results_df.to_excel(writer, index=True, sheet_name='Analysis Results')

            analysis_worksheet = writer.sheets['Analysis Results']
            for row in analysis_worksheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border
                    cell.alignment = alignment
            for cell in analysis_worksheet[1]:
                cell.font = bold_font
            for cell in analysis_worksheet['A']:
                cell.font = bold_font

        output.seek(0)
        
        return send_file(output, download_name="Model_Analysis_Results.xlsx", as_attachment=True)
    else:
        return 'Invalid file format. Please upload an Excel file.'
    
def calculate_results(df, subjects):
    combined_results = {}
    for subject in subjects:
        combined_results[subject] = {
            'Absent': 0,
            '0-15': 0,
            '16-30': 0,
            '31-44': 0,
            '45-65': 0,
            '66-85': 0,
            '86-100': 0,
            'Total Failures': 0,
            'Total Passes': 0,
            'Pass %': 0,
            'University All Clear Model Fail': 0,
            'University All Clear Model Pass': 0,
            'University Arrear Model Fail': 0,
            'University Arrear Model Pass': 0
        }

    failure_counts = []

    for index, row in df.iterrows():
        fail_count = 0
        arrears = int(row.get('ARREAR', 0))

        for subject in subjects:
            marks = row.get(subject, None)
            if marks is None:
                continue

            if marks == 'AB':
                combined_results[subject]['Absent'] += 1
                continue

            try:
                marks = int(marks)
            except ValueError:
                continue

            if marks < 45:
                fail_count += 1

            if marks <= 15:
                combined_results[subject]['0-15'] += 1
            elif marks <= 30:
                combined_results[subject]['16-30'] += 1
            elif marks <= 44:
                combined_results[subject]['31-44'] += 1
            elif marks <= 65:
                combined_results[subject]['45-65'] += 1
            elif marks <= 85:
                combined_results[subject]['66-85'] += 1
            else:
                combined_results[subject]['86-100'] += 1

            if marks < 45:
                combined_results[subject]['Total Failures'] += 1
            else:
                combined_results[subject]['Total Passes'] += 1

            if marks < 45:
                if arrears == 0:
                    combined_results[subject]['University All Clear Model Fail'] += 1
                else:
                    combined_results[subject]['University Arrear Model Fail'] += 1
            elif marks > 44:
                if arrears == 0:
                    combined_results[subject]['University All Clear Model Pass'] += 1
                else:
                    combined_results[subject]['University Arrear Model Pass'] += 1

        failure_counts.append(fail_count)

    for subject in subjects:
        stats = combined_results[subject]
        total_students = stats['Total Failures'] + stats['Total Passes']
        stats['Pass %'] = round((stats['Total Passes'] / total_students) * 100, 2) if total_students > 0 else 0.00

    return combined_results, failure_counts

if __name__ == '__main__':
    app.run(debug=True)
